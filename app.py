import streamlit as st
import pandas as pd
import pyreadstat
import io
import tempfile
import os
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide", page_title="Survey Productivity Dashboard", page_icon="📋")

# ── Theme toggle (must be first sidebar widget) ──────────────────────────────
with st.sidebar:
    st.markdown("### 🎨 Appearance")
    theme_choice = st.radio("Theme", ["Dark", "Light"], horizontal=True, label_visibility="collapsed")

DARK   = theme_choice == "Dark"
ACCENT = "#c01823"

if DARK:
    BG, BG2, BORDER       = "#0d0f14", "#111318", "#1e2230"
    TEXT, TEXT_MUTED       = "#e0e6f0", "#4a5a7a"
    TEXT_SUB               = "#8a9bc4"
    CHART_BG, CHART_PLOT   = "#0d0f14", "#111318"
    CHART_GRID, CHART_TEXT = "#1e2230", "#8a9bc4"
    CELL_ZERO = "background-color:#1a1a1a; color:#333;"
    CELL_LOW  = "background-color:#2a1500; color:#f5a623;"
    CELL_HIGH = "background-color:#0a200a; color:#4caf50;"
    WARN_BG, WARN_FG       = "#1a1208", "#f5a623"
    INFO_BG,  INFO_FG      = "#0a1020", "#c06070"
else:
    BG, BG2, BORDER       = "#f4f5f7", "#ffffff", "#dde1ea"
    TEXT, TEXT_MUTED       = "#1a1e2e", "#7a8299"
    TEXT_SUB               = "#5a6380"
    CHART_BG, CHART_PLOT   = "#f4f5f7", "#ffffff"
    CHART_GRID, CHART_TEXT = "#e0e4ed", "#5a6380"
    CELL_ZERO = "background-color:#f0f0f0; color:#bbb;"
    CELL_LOW  = "background-color:#fff3cd; color:#856404;"
    CELL_HIGH = "background-color:#d4edda; color:#155724;"
    WARN_BG, WARN_FG       = "#fffbf0", "#856404"
    INFO_BG,  INFO_FG      = "#f0f4ff", ACCENT

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html,body,[class*="css"]{{font-family:'IBM Plex Sans',sans-serif;}}
.stApp{{background-color:{BG};color:{TEXT};}}
[data-testid="stSidebar"]{{background-color:{BG2};border-right:1px solid {BORDER};}}
[data-testid="stSidebar"] label,[data-testid="stSidebar"] p{{color:{TEXT_SUB}!important;font-size:.75rem;text-transform:uppercase;letter-spacing:.08em;}}
.main-title{{font-family:'IBM Plex Mono',monospace;font-size:1.6rem;font-weight:600;color:{TEXT};border-bottom:2px solid {ACCENT};padding-bottom:.4rem;margin-bottom:.2rem;}}
.main-subtitle{{font-size:.78rem;color:{TEXT_MUTED};font-family:'IBM Plex Mono',monospace;letter-spacing:.05em;margin-bottom:1.5rem;}}
.metric-row{{display:flex;gap:1rem;margin-bottom:1.5rem;flex-wrap:wrap;}}
.metric-card{{background:{BG2};border:1px solid {BORDER};border-top:3px solid {ACCENT};border-radius:4px;padding:1rem 1.2rem;flex:1;min-width:130px;}}
.metric-card .label{{font-family:'IBM Plex Mono',monospace;font-size:.65rem;color:{TEXT_MUTED};text-transform:uppercase;letter-spacing:.1em;margin-bottom:.3rem;}}
.metric-card .value{{font-family:'IBM Plex Mono',monospace;font-size:1.6rem;font-weight:600;color:{ACCENT};}}
.metric-card .sub{{font-size:.7rem;color:{TEXT_MUTED};margin-top:.2rem;}}
.section-header{{font-family:'IBM Plex Mono',monospace;font-size:.7rem;text-transform:uppercase;letter-spacing:.15em;color:{TEXT_MUTED};border-bottom:1px solid {BORDER};padding-bottom:.4rem;margin:1.5rem 0 1rem 0;}}
[data-testid="stDataFrame"]{{border:1px solid {BORDER};border-radius:4px;}}
.warning-box{{background:{WARN_BG};border-left:3px solid #f5a623;padding:.6rem 1rem;border-radius:0 4px 4px 0;font-size:.8rem;color:{WARN_FG};margin:.5rem 0;font-family:'IBM Plex Mono',monospace;}}
.info-box{{background:{INFO_BG};border-left:3px solid {ACCENT};padding:.6rem 1rem;border-radius:0 4px 4px 0;font-size:.8rem;color:{INFO_FG};margin:.5rem 0;font-family:'IBM Plex Mono',monospace;}}
.stDownloadButton>button{{background-color:{ACCENT};color:white;border:none;border-radius:3px;padding:.5rem 1.4rem;font-family:'IBM Plex Mono',monospace;font-size:.75rem;letter-spacing:.05em;text-transform:uppercase;}}
.stDownloadButton>button:hover{{background-color:#9a1018;}}
.stTabs [data-baseweb="tab-list"]{{background:{BG2};border-bottom:1px solid {BORDER};gap:0;}}
.stTabs [data-baseweb="tab"]{{font-family:'IBM Plex Mono',monospace;font-size:.7rem;text-transform:uppercase;letter-spacing:.1em;color:{TEXT_MUTED};border-radius:0;padding:.6rem 1.2rem;}}
.stTabs [aria-selected="true"]{{color:{ACCENT}!important;border-bottom:2px solid {ACCENT};background:transparent!important;}}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">Survey Productivity Dashboard</div>', unsafe_allow_html=True)
st.markdown('<div class="main-subtitle">Enumerator performance tracking & quality control</div>', unsafe_allow_html=True)

with st.sidebar:
    st.markdown("---")
    st.markdown("### 📂 Data Source")
    uploaded_file = st.file_uploader("Upload File", type=["dta","xlsx","xls"])
    st.markdown("---")
    st.markdown("### 🗓 Date Format")
    header_style = st.selectbox("Date Header Style",
        ["Pretty (e.g., 10 Sep 2025)","Safe (e.g., d_10Sep2025)","Compact (e.g., 10Sep2025)","ISO (e.g., 2025-09-10)"],
        label_visibility="collapsed")
    st.markdown("---")
    st.markdown("### ⚙️ Filters")
    date_range_filter = st.checkbox("Filter by date range", value=False)
    st.markdown("---")
    st.markdown("### 🚨 QC Thresholds")
    outlier_low  = st.slider("Flag if daily count below", 0, 10, 2)
    outlier_high = st.slider("Flag if daily count above", 5, 100, 20)

if uploaded_file is None:
    st.markdown('<div class="info-box">⬅ Upload a .dta, .xlsx, or .xls file from the sidebar to get started.</div>', unsafe_allow_html=True)
    st.stop()

# ── Read file ────────────────────────────────────────────────────────────────
try:
    file_bytes = uploaded_file.read()
    if uploaded_file.name.lower().endswith(".dta"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".dta") as tmp:
            tmp.write(file_bytes); tmp_path = tmp.name
        df, meta = pyreadstat.read_dta(tmp_path, apply_value_formats=True)
        os.unlink(tmp_path)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    with st.sidebar:
        st.success(f"✓ {len(df):,} rows loaded")
except Exception as e:
    st.error(f"Could not read file: {e}"); st.stop()

if isinstance(df.columns, pd.MultiIndex):
    df.columns = ["_".join(map(str,c)) for c in df.columns]
df.columns = [str(c).strip() for c in df.columns]

col_options = ["Select a column"] + list(df.columns)

with st.sidebar:
    st.markdown("---")
    st.markdown("### 🔗 Column Mapping")
    enum_col    = st.selectbox("Enumerator Column",           col_options, index=col_options.index("enum")      if "enum"      in col_options else 0)
    date_col    = st.selectbox("Date Column",                 col_options, index=col_options.index("starttime") if "starttime" in col_options else 0)
    consent_col = st.selectbox("Consent Column (optional)",   col_options)
    addr1_col   = st.selectbox("Address Column 1 (optional)", col_options)
    addr2_col   = st.selectbox("Address Column 2 (optional)", col_options)
    uid_col     = st.selectbox("Unique ID Column (optional)",  col_options)

if enum_col == "Select a column" or date_col == "Select a column":
    st.markdown('<div class="warning-box">⚠ Please map the Enumerator and Date columns in the sidebar.</div>', unsafe_allow_html=True)
    st.stop()

rename_map = {enum_col:"enum", date_col:"date"}
if consent_col != "Select a column": rename_map[consent_col] = "consent"
if addr1_col   != "Select a column": rename_map[addr1_col]   = "addr1"
if addr2_col   != "Select a column": rename_map[addr2_col]   = "addr2"
if uid_col     != "Select a column": rename_map[uid_col]     = "uid"
df = df.rename(columns=rename_map)

df["date"] = pd.to_datetime(df["date"], errors="coerce")
if hasattr(df["date"].dtype,"tz") and df["date"].dtype.tz is not None:
    df["date"] = df["date"].dt.tz_localize(None)
df = df.dropna(subset=["date"])
df["date"] = df["date"].dt.normalize()

def clean(x): return "Unknown" if pd.isna(x) else str(x).strip()
df["enum"] = df["enum"].apply(clean)
for c in ["addr1","addr2"]:
    if c in df.columns: df[c] = df[c].apply(clean)

if "consent" in df.columns:
    df["Consent_Status"] = df["consent"].apply(
        lambda x: "Yes" if str(x).lower().strip() in ["yes","1","true","y"] else "No")

duplicate_warnings = []
if "uid" in df.columns:
    dups = df[df.duplicated(subset=["uid"], keep=False)]
    if not dups.empty:
        duplicate_warnings.append(f"⚠ {dups['uid'].nunique()} unique IDs appear more than once (possible duplicates).")

min_date = df["date"].min().date()
max_date = df["date"].max().date()

if date_range_filter:
    with st.sidebar:
        st.markdown("**Date Range**")
        date_from = st.date_input("From", value=min_date, min_value=min_date, max_value=max_date)
        date_to   = st.date_input("To",   value=max_date, min_value=min_date, max_value=max_date)
    df = df[(df["date"].dt.date >= date_from) & (df["date"].dt.date <= date_to)]
    if df.empty: st.warning("No data in selected date range."); st.stop()

with st.sidebar:
    all_enums      = sorted(df["enum"].unique().tolist())
    selected_enums = st.multiselect("Filter Enumerators", all_enums, default=all_enums)
if selected_enums: df = df[df["enum"].isin(selected_enums)]
if df.empty: st.warning("No data after filtering."); st.stop()

total_surveys = len(df)
total_enums   = df["enum"].nunique()
total_days    = df["date"].nunique()
avg_per_day   = total_surveys / total_days if total_days > 0 else 0
consent_pct   = (df["Consent_Status"]=="Yes").mean()*100 if "Consent_Status" in df.columns else None

consent_card = (
    f"<div class='metric-card'><div class='label'>Consent Rate</div>"
    f"<div class='value'>{consent_pct:.1f}%</div><div class='sub'>respondents consented</div></div>"
) if consent_pct is not None else ""

st.markdown(f"""
<div class="metric-row">
  <div class="metric-card"><div class="label">Total Surveys</div><div class="value">{total_surveys:,}</div><div class="sub">{min_date.strftime('%d %b')} → {max_date.strftime('%d %b %Y')}</div></div>
  <div class="metric-card"><div class="label">Enumerators</div><div class="value">{total_enums}</div><div class="sub">active in period</div></div>
  <div class="metric-card"><div class="label">Field Days</div><div class="value">{total_days}</div><div class="sub">days with submissions</div></div>
  <div class="metric-card"><div class="label">Avg / Day</div><div class="value">{avg_per_day:.1f}</div><div class="sub">surveys per day (all)</div></div>
  {consent_card}
</div>
""", unsafe_allow_html=True)

for w in duplicate_warnings:
    st.markdown(f'<div class="warning-box">{w}</div>', unsafe_allow_html=True)

group_cols = ["enum"]
if "addr1"          in df.columns: group_cols.append("addr1")
if "addr2"          in df.columns: group_cols.append("addr2")
if "Consent_Status" in df.columns: group_cols.append("Consent_Status")

daily = (df.groupby(group_cols+["date"], dropna=False, observed=True)
           .size().reset_index(name="count"))

pivot = (daily.pivot_table(index=group_cols, columns="date", values="count",
                            fill_value=0, aggfunc="sum", observed=True)
               .reset_index())

new_cols, seen = [], {}
for col in pivot.columns:
    base = "d_"+col.strftime("%d%b%Y") if isinstance(col, pd.Timestamp) else str(col)
    if base in seen: seen[base]+=1; base=f"{base}_{seen[base]}"
    else: seen[base]=0
    new_cols.append(base)
pivot.columns = new_cols
date_cols = [c for c in pivot.columns if c.startswith("d_")]

pivot["Total"]       = pivot[date_cols].sum(axis=1).astype(int)
pivot["Avg/Day"]     = pivot[date_cols].replace(0, np.nan).mean(axis=1).round(1)
pivot["Days Active"] = (pivot[date_cols]>0).sum(axis=1).astype(int)

def flag_enum(row):
    active = row[date_cols]; active = active[active>0]; flags=[]
    if any(active<outlier_low):  flags.append(f"Low day (<{outlier_low})")
    if any(active>outlier_high): flags.append(f"High day (>{outlier_high})")
    return ", ".join(flags) if flags else "✓ OK"
pivot["QC Flag"] = pivot.apply(flag_enum, axis=1)

pretty = pivot.copy()
rename_pretty, seen_pretty = {}, {}
for col in pretty.columns:
    if col.startswith("d_"):
        raw = col[2:]
        try:
            dt = datetime.strptime(raw, "%d%b%Y")
            if   header_style.startswith("Pretty"):  name = dt.strftime("%d %b %Y")
            elif header_style.startswith("Compact"): name = dt.strftime("%d%b%Y")
            elif header_style.startswith("ISO"):      name = dt.strftime("%Y-%m-%d")
            else:                                      name = col
        except: name = col
        if name in seen_pretty: seen_pretty[name]+=1; name=f"{name}_{seen_pretty[name]}"
        else: seen_pretty[name]=0
        rename_pretty[col] = name
pretty.rename(columns=rename_pretty, inplace=True)
pretty = pretty.loc[:, ~pretty.columns.duplicated()]

daily_total_row = {}
for col in pretty.columns:
    if col in group_cols:        daily_total_row[col] = "── TOTAL ──" if col=="enum" else ""
    elif col=="Total":           daily_total_row[col] = int(pretty[col].sum())
    elif col=="Days Active":     daily_total_row[col] = int(pretty[col].sum())
    elif col=="Avg/Day":         daily_total_row[col] = round(float(pretty[col].mean()),1)
    elif col=="QC Flag":         daily_total_row[col] = "──"
    else:
        try:    daily_total_row[col] = pretty[col].sum()
        except: daily_total_row[col] = ""

pretty_with_totals = pd.concat([pretty, pd.DataFrame([daily_total_row])], ignore_index=True)
pretty_with_totals["Avg/Day"] = pd.to_numeric(pretty_with_totals["Avg/Day"], errors="coerce")

tab1, tab2, tab3, tab4 = st.tabs(["📋 Productivity Table","📊 Charts","🚨 QC Report","📥 Export"])

# ══ TAB 1 ══════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="section-header">Daily Survey Counts by Enumerator</div>', unsafe_allow_html=True)
    date_display_cols = list(rename_pretty.values())
    _num = pretty_with_totals[[c for c in date_display_cols if c in pretty_with_totals.columns]].apply(pd.to_numeric, errors="coerce")
    _max = _num.max().max() or 1

    def heatmap_style(val):
        if not isinstance(val,(int,float)) or pd.isna(val): return ""
        if val==0:            return CELL_ZERO
        if val<outlier_low:   return CELL_LOW
        if val>outlier_high:  return CELL_HIGH
        intensity = min(val/_max, 1.0)
        if DARK:
            r = int(20  + (192-20)*intensity)
            g = int(5   + (24-5)*intensity)
            b = int(5   + (35-5)*intensity)
            tc = "#e0e6f0" if intensity>0.3 else "#888"
        else:
            r = int(255-(255-192)*intensity)
            g = int(255-(255-24)*intensity)
            b = int(255-(255-35)*intensity)
            tc = "#ffffff" if intensity>0.55 else "#1a1e2e"
        return f"background-color:rgb({r},{g},{b});color:{tc};"

    def qc_style(val):
        if val=="✓ OK": return f"color:{'#4caf50' if DARK else '#1a7a3a'};font-weight:600;"
        return "color:#f5a623;font-weight:600;"

    styled = (pretty_with_totals.style
        .map(heatmap_style, subset=[c for c in date_display_cols if c in pretty_with_totals.columns])
        .map(qc_style,      subset=["QC Flag"] if "QC Flag" in pretty_with_totals.columns else [])
        .set_properties(**{"font-family":"IBM Plex Mono,monospace","font-size":"12px"}))
    st.dataframe(styled, width="stretch", height=500)
    st.caption(f"🔴 Normal (red gradient)  🟠 Below threshold (<{outlier_low})  🟢 Above threshold (>{outlier_high})  ⬛ Zero")

# ══ TAB 2 ══════════════════════════════════════════════════════════════════════
with tab2:
    import plotly.graph_objects as go
    ACCENT2 = "#e8606a"

    def apply_theme(fig):
        fig.update_layout(
            paper_bgcolor=CHART_BG, plot_bgcolor=CHART_PLOT,
            font=dict(family="IBM Plex Mono",color=CHART_TEXT,size=11),
            xaxis=dict(gridcolor=CHART_GRID,zerolinecolor=CHART_GRID),
            yaxis=dict(gridcolor=CHART_GRID,zerolinecolor=CHART_GRID),
            margin=dict(l=40,r=20,t=40,b=40))
        return fig

    ca, cb = st.columns(2)
    with ca:
        st.markdown('<div class="section-header">Total Surveys per Enumerator</div>', unsafe_allow_html=True)
        bdf = pretty[["enum","Total"]].sort_values("Total",ascending=True).tail(20)
        fig=go.Figure(go.Bar(x=bdf["Total"],y=bdf["enum"],orientation="h",marker=dict(color=ACCENT,line=dict(width=0))))
        apply_theme(fig); fig.update_layout(height=400,showlegend=False)
        st.plotly_chart(fig, width="stretch")
    with cb:
        st.markdown('<div class="section-header">Average Surveys / Active Day</div>', unsafe_allow_html=True)
        adf=pretty[["enum","Avg/Day"]].dropna().sort_values("Avg/Day",ascending=True).tail(20)
        fig2=go.Figure(go.Bar(x=adf["Avg/Day"],y=adf["enum"],orientation="h",marker=dict(color=ACCENT2,line=dict(width=0))))
        apply_theme(fig2); fig2.update_layout(height=400,showlegend=False)
        st.plotly_chart(fig2, width="stretch")

    st.markdown('<div class="section-header">Daily Submission Volume</div>', unsafe_allow_html=True)
    dt2 = daily.groupby("date")["count"].sum().reset_index()
    dt2.columns=["date","count"]
    fig3=go.Figure()
    fig3.add_trace(go.Scatter(x=dt2["date"],y=dt2["count"],mode="lines+markers",
        line=dict(color=ACCENT,width=2),marker=dict(size=6,color=ACCENT),
        fill="tozeroy",fillcolor="rgba(192,24,35,0.10)",name="Daily Total"))
    if len(dt2)>=3:
        dt2["rolling"]=dt2["count"].rolling(3,center=True).mean()
        fig3.add_trace(go.Scatter(x=dt2["date"],y=dt2["rolling"],mode="lines",
            line=dict(color="#f5a623",width=1.5,dash="dot"),name="3-Day Avg"))
    apply_theme(fig3); fig3.update_layout(height=300,legend=dict(orientation="h",yanchor="bottom",y=1.02))
    st.plotly_chart(fig3, width="stretch")

    st.markdown('<div class="section-header">Enumerator Activity Heatmap</div>', unsafe_allow_html=True)
    hd = pivot[["enum"]+date_cols].set_index("enum")
    hd.columns=[c[2:] for c in hd.columns]
    cs = ([[0,CHART_PLOT],[0.001,"#2a0a0a"],[0.4,"#7a1018"],[1,ACCENT]] if DARK
          else [[0,"#ffffff"],[0.001,"#fde8ea"],[0.4,"#e87080"],[1,ACCENT]])
    fig4=go.Figure(go.Heatmap(z=hd.values,x=hd.columns.tolist(),y=hd.index.tolist(),
        colorscale=cs,showscale=True,hoverongaps=False))
    apply_theme(fig4); fig4.update_layout(height=max(300,len(hd)*22+80))
    st.plotly_chart(fig4, width="stretch")

    if "Consent_Status" in df.columns:
        st.markdown('<div class="section-header">Consent Breakdown</div>', unsafe_allow_html=True)
        cc=df["Consent_Status"].value_counts()
        fig5=go.Figure(go.Pie(labels=cc.index.tolist(),values=cc.values.tolist(),hole=0.55,
            marker=dict(colors=[ACCENT,"#f5a623"],line=dict(color=CHART_BG,width=3)),
            textfont=dict(family="IBM Plex Mono",size=11)))
        apply_theme(fig5); fig5.update_layout(height=320,showlegend=True,
            legend=dict(orientation="h",yanchor="bottom",y=-0.15))
        st.plotly_chart(fig5, width="stretch")

# ══ TAB 3 ══════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="section-header">Quality Control Flags</div>', unsafe_allow_html=True)
    flagged=pretty[pretty["QC Flag"]!="✓ OK"][["enum","Total","Avg/Day","Days Active","QC Flag"]]
    if flagged.empty:
        st.markdown('<div class="info-box">✓ No QC flags raised. All enumerators within thresholds.</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="warning-box">⚠ {len(flagged)} enumerator(s) have QC flags.</div>', unsafe_allow_html=True)
        st.dataframe(flagged, width="stretch")
    st.markdown('<div class="section-header">Enumerator Summary Statistics</div>', unsafe_allow_html=True)
    summary=pretty[["enum","Total","Avg/Day","Days Active","QC Flag"]].sort_values("Total",ascending=False)
    st.dataframe(summary, width="stretch", height=400)
    if duplicate_warnings:
        st.markdown('<div class="section-header">Duplicate Warnings</div>', unsafe_allow_html=True)
        for w in duplicate_warnings: st.markdown(f'<div class="warning-box">{w}</div>', unsafe_allow_html=True)
        if "uid" in df.columns:
            dup_df=df[df.duplicated(subset=["uid"],keep=False)].sort_values("uid")
            sc=["uid","enum","date"]+( ["addr1"] if "addr1" in df.columns else [])
            st.dataframe(dup_df[sc], width="stretch", height=300)
    st.markdown('<div class="section-header">Coverage Gap Analysis</div>', unsafe_allow_html=True)
    all_dates=pd.date_range(df["date"].min(),df["date"].max(),freq="D")
    missing_dates=[d for d in all_dates if d not in df["date"].unique()]
    if missing_dates:
        st.markdown(f'<div class="warning-box">⚠ No submissions on: {", ".join(d.strftime("%d %b") for d in missing_dates)}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="info-box">✓ Submissions recorded every day in the date range.</div>', unsafe_allow_html=True)

# ══ TAB 4 ══════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="section-header">Export Options</div>', unsafe_allow_html=True)
    c1,c2=st.columns(2)

    with c1:
        st.markdown("**📊 Formatted Excel**")
        st.caption("No fill for zeros. Light grey → near-black font gradient for normal values. Amber / green tints for QC flags.")

        out=io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            pretty_with_totals.to_excel(writer, index=False, sheet_name="Productivity")
            summary.to_excel(writer, index=False, sheet_name="Summary")
            ws=writer.sheets["Productivity"]

            # Header
            hfill=PatternFill("solid",fgColor="1A1A1A")
            hfont=Font(name="Garamond",bold=True,color="FFFFFF",size=11)
            hbdr =Border(bottom=Side(style="medium",color="C01823"))
            for cell in ws[1]:
                cell.fill=hfill; cell.font=hfont
                cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border=hbdr

            date_display_cols_set = set(date_display_cols)
            date_col_idx = {i+1 for i,col in enumerate(pretty_with_totals.columns) if col in date_display_cols_set}

            # Max value for gradient
            nvals=[]
            for row in ws.iter_rows(min_row=2,max_row=ws.max_row-1):
                for cell in row:
                    if cell.column in date_col_idx and isinstance(cell.value,(int,float)):
                        nvals.append(cell.value)
            max_val=max(nvals) if nvals else 1

            FONT_NORMAL = Font(name="Garamond", size=11, color="222222")
            FONT_ZERO   = Font(name="Garamond", size=11, color="CCCCCC")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1):
                for cell in row:
                    if cell.column in date_col_idx:
                        cell.font = FONT_ZERO if (cell.value is None or cell.value == 0) else FONT_NORMAL
                    else:
                        cell.font = FONT_NORMAL
                    cell.alignment = Alignment(horizontal="center")

            # Totals row
            tf=PatternFill("solid",fgColor="F2F2F2")
            for cell in ws[ws.max_row]:
                cell.fill=tf
                cell.font=Font(name="Garamond",bold=True,color="333333",size=11)
                cell.border=Border(top=Side(style="thin",color="CCCCCC"))
                cell.alignment=Alignment(horizontal="center")

            for col_cells in ws.columns:
                ml=max((len(str(c.value or "")) for c in col_cells),default=0)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width=min(ml+2,20)
            ws.freeze_panes="D2"

        st.download_button("⬇ Download Excel", out.getvalue(),
            file_name=f"productivity_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with c2:
        st.markdown("**📄 CSV Export**")
        st.caption("Plain CSV, no formatting.")
        st.download_button("⬇ Download CSV",
            pretty_with_totals.to_csv(index=False).encode("utf-8"),
            file_name=f"productivity_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv")

    st.markdown("---")
    st.markdown('<div class="section-header">QC Report Export</div>', unsafe_allow_html=True)
    st.download_button("⬇ Download QC Report (CSV)",
        summary.to_csv(index=False).encode("utf-8"),
        file_name=f"qc_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv")

    st.markdown("---")
    st.caption(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Rows: {total_surveys:,} · Enumerators: {total_enums} · Period: {min_date} → {max_date}")
